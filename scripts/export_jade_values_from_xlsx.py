#!/usr/bin/env python3
"""Export jade recipe values and ranking table data from data/金符文菜玉璧.xlsx.

Default behavior:
1. Read workbook: <cwd>/data/金符文菜玉璧.xlsx
2. Read active sheet
3. Starting from row 3, use column B as recipe name and column C as jade value
4. Overwrite data/jade-values.json
5. Overwrite data/jade-ranking-table.json for the ranking modal

Examples:
    python3 scripts/export_jade_values_from_xlsx.py
    python3 scripts/export_jade_values_from_xlsx.py --sheet Sheet1
    python3 scripts/export_jade_values_from_xlsx.py --input /path/to/金符文菜玉璧.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


DEFAULT_INPUT = "data/金符文菜玉璧.xlsx"
DEFAULT_START_ROW = 3
DEFAULT_NAME_COLUMN = "B"
DEFAULT_VALUE_COLUMN = "C"
DEFAULT_VALUES_OUTPUT = "data/jade-values.json"
DEFAULT_RANKING_OUTPUT = "data/jade-ranking-table.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read recipe jade values from an xlsx file and print jade-values.json content."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help=f"Workbook path. Default: {DEFAULT_INPUT}",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Sheet name to read. Default: active sheet.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=DEFAULT_START_ROW,
        help=f"First data row. Default: {DEFAULT_START_ROW}",
    )
    parser.add_argument(
        "--name-column",
        default=DEFAULT_NAME_COLUMN,
        help=f"Recipe name column. Default: {DEFAULT_NAME_COLUMN}",
    )
    parser.add_argument(
        "--value-column",
        default=DEFAULT_VALUE_COLUMN,
        help=f"Jade value column. Default: {DEFAULT_VALUE_COLUMN}",
    )
    parser.add_argument(
        "--values-output",
        default=DEFAULT_VALUES_OUTPUT,
        help=f"jade-values.json output path. Default: {DEFAULT_VALUES_OUTPUT}",
    )
    parser.add_argument(
        "--ranking-output",
        default=DEFAULT_RANKING_OUTPUT,
        help=f"Ranking table json output path. Default: {DEFAULT_RANKING_OUTPUT}",
    )
    return parser.parse_args()


def normalize_recipe_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_jade_value(value: Any) -> int | float | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        numeric = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            numeric = float(text)
        except ValueError:
            return None

    if numeric.is_integer():
        return int(numeric)
    return numeric


def load_workbook_sheet(input_path: Path, sheet_name: str):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install it first, e.g. `python3 -m pip install openpyxl`.") from exc

    workbook = load_workbook(input_path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(
                "Sheet not found: {}. Available sheets: {}".format(
                    sheet_name, ", ".join(workbook.sheetnames)
                )
            )
        return workbook[sheet_name]
    return workbook.active


def normalize_cell_display(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "hour") and hasattr(value, "minute") and hasattr(value, "second"):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second + getattr(value, "microsecond", 0) / 1_000_000
        minutes = int(total_seconds // 60)
        seconds = total_seconds - minutes * 60
        if abs(seconds - round(seconds)) < 1e-9:
            return f"{minutes:02d}分{int(round(seconds)):02d}秒"
        return f"{minutes:02d}分{seconds:04.1f}秒"
    return str(value).strip()


def extract_values(sheet, start_row: int, name_column: str, value_column: str) -> dict[str, int | float]:
    result: dict[str, int | float] = {}

    row = start_row
    while True:
        recipe_name = normalize_recipe_name(sheet[f"{name_column}{row}"].value)
        jade_value = normalize_jade_value(sheet[f"{value_column}{row}"].value)

        if not recipe_name and jade_value is None:
            break

        if recipe_name:
            if jade_value is None:
                raise RuntimeError(f"Row {row}: recipe '{recipe_name}' has an empty or invalid jade value.")
            result[recipe_name] = jade_value

        row += 1

    return result


def extract_ranking_table(sheet) -> dict[str, Any]:
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row < 2:
        raise RuntimeError("Ranking sheet does not contain enough rows.")

    title = normalize_cell_display(sheet["A1"].value)
    headers = [normalize_cell_display(sheet.cell(2, col).value) for col in range(1, max_col + 1)]
    rows: list[list[str]] = []

    for row_idx in range(3, max_row + 1):
        row_values = [normalize_cell_display(sheet.cell(row_idx, col).value) for col in range(1, max_col + 1)]
        if not any(row_values):
            continue
        rows.append(row_values)

    return {
        "title": title,
        "headers": headers,
        "rows": rows,
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = Path.cwd() / input_path
    values_output_path = Path(args.values_output)
    if not values_output_path.is_absolute():
        values_output_path = Path.cwd() / values_output_path
    ranking_output_path = Path(args.ranking_output)
    if not ranking_output_path.is_absolute():
        ranking_output_path = Path.cwd() / ranking_output_path

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    try:
        sheet = load_workbook_sheet(input_path, args.sheet)
        values = extract_values(
            sheet=sheet,
            start_row=args.start_row,
            name_column=args.name_column.upper(),
            value_column=args.value_column.upper(),
        )
        ranking_table = extract_ranking_table(sheet)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    write_json(values_output_path, values)
    write_json(ranking_output_path, ranking_table)
    print(f"updated: {values_output_path}")
    print(f"updated: {ranking_output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
